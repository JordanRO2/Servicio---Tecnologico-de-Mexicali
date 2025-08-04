"""
Sistema de Generación Automatizada de Reportes de Indicadores
Departamento de Sistemas y Computación - ITM
Autor: Jordan Rivera Rodriguez
Fecha: Junio 2025

Este script automatiza la generación de reportes mensuales de indicadores del departamento,
procesando datos de múltiples fuentes y generando visualizaciones interactivas.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

class GeneradorReportesIndicadores:
    """
    Clase principal para la generación automatizada de reportes de indicadores departamentales.
    
    Attributes:
        ruta_datos (str): Ruta al directorio con los archivos de datos fuente
        ruta_salida (str): Ruta donde se guardarán los reportes generados
    """
    
    def __init__(self, ruta_datos='./datos', ruta_salida='./reportes'):
        """
        Inicializa el generador de reportes.
        
        Args:
            ruta_datos (str): Directorio con archivos de entrada
            ruta_salida (str): Directorio para reportes generados
        """
        self.ruta_datos = ruta_datos
        self.ruta_salida = ruta_salida
        self._crear_directorios()
        
    def _crear_directorios(self):
        """Crea los directorios necesarios si no existen."""
        for ruta in [self.ruta_datos, self.ruta_salida]:
            if not os.path.exists(ruta):
                os.makedirs(ruta)
                
    def cargar_datos_incidencias(self, archivo='incidencias.csv'):
        """
        Carga y procesa los datos de incidencias técnicas.
        
        Args:
            archivo (str): Nombre del archivo CSV con datos de incidencias
            
        Returns:
            pd.DataFrame: DataFrame con datos procesados de incidencias
        """
        try:
            df = pd.read_csv(os.path.join(self.ruta_datos, archivo), 
                           parse_dates=['fecha_reporte', 'fecha_resolucion'])
            
            # Calcular tiempo de resolución
            df['tiempo_resolucion_horas'] = (
                df['fecha_resolucion'] - df['fecha_reporte']
            ).dt.total_seconds() / 3600
            
            # Categorizar por prioridad
            df['categoria_tiempo'] = pd.cut(
                df['tiempo_resolucion_horas'],
                bins=[0, 4, 24, 72, float('inf')],
                labels=['Inmediata', 'Rápida', 'Normal', 'Demorada']
            )
            
            return df
            
        except Exception as e:
            print(f"Error al cargar datos de incidencias: {e}")
            return pd.DataFrame()
    
    def generar_estadisticas_mensuales(self, df_incidencias, mes, año):
        """
        Genera estadísticas mensuales de incidencias.
        
        Args:
            df_incidencias (pd.DataFrame): DataFrame con datos de incidencias
            mes (int): Mes a analizar
            año (int): Año a analizar
            
        Returns:
            dict: Diccionario con estadísticas calculadas
        """
        # Filtrar por mes y año
        df_mes = df_incidencias[
            (df_incidencias['fecha_reporte'].dt.month == mes) &
            (df_incidencias['fecha_reporte'].dt.year == año)
        ]
        
        estadisticas = {
            'total_incidencias': len(df_mes),
            'incidencias_resueltas': len(df_mes[df_mes['estado'] == 'Resuelto']),
            'tiempo_promedio_resolucion': df_mes['tiempo_resolucion_horas'].mean(),
            'incidencias_por_tipo': df_mes['tipo_incidencia'].value_counts().to_dict(),
            'incidencias_por_prioridad': df_mes['prioridad'].value_counts().to_dict(),
            'tasa_resolucion': (
                len(df_mes[df_mes['estado'] == 'Resuelto']) / len(df_mes) * 100 
                if len(df_mes) > 0 else 0
            )
        }
        
        return estadisticas
    
    def crear_visualizaciones(self, estadisticas, mes, año):
        """
        Crea visualizaciones para el reporte mensual.
        
        Args:
            estadisticas (dict): Estadísticas calculadas
            mes (int): Mes del reporte
            año (int): Año del reporte
            
        Returns:
            list: Lista de rutas a las imágenes generadas
        """
        plt.style.use('seaborn-v0_8-darkgrid')
        imagenes = []
        
        # Gráfico 1: Distribución por tipo de incidencia
        if estadisticas['incidencias_por_tipo']:
            fig, ax = plt.subplots(figsize=(10, 6))
            tipos = list(estadisticas['incidencias_por_tipo'].keys())
            valores = list(estadisticas['incidencias_por_tipo'].values())
            
            bars = ax.bar(tipos, valores, color='skyblue', edgecolor='navy')
            ax.set_title(f'Incidencias por Tipo - {mes}/{año}', fontsize=16, fontweight='bold')
            ax.set_xlabel('Tipo de Incidencia', fontsize=12)
            ax.set_ylabel('Cantidad', fontsize=12)
            
            # Agregar valores en las barras
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom')
            
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            
            ruta_img = os.path.join(self.ruta_salida, f'incidencias_tipo_{mes}_{año}.png')
            plt.savefig(ruta_img, dpi=300, bbox_inches='tight')
            plt.close()
            imagenes.append(ruta_img)
        
        # Gráfico 2: Tasa de resolución
        fig, ax = plt.subplots(figsize=(8, 8))
        tasa = estadisticas['tasa_resolucion']
        
        # Crear gráfico de dona
        sizes = [tasa, 100 - tasa]
        colors = ['#2ecc71', '#e74c3c']
        labels = ['Resueltas', 'Pendientes']
        
        wedges, texts, autotexts = ax.pie(sizes, labels=labels, colors=colors,
                                          autopct='%1.1f%%', startangle=90,
                                          pctdistance=0.85)
        
        # Crear círculo interior para efecto dona
        centre_circle = plt.Circle((0, 0), 0.70, fc='white')
        fig.gca().add_artist(centre_circle)
        
        ax.set_title(f'Tasa de Resolución - {mes}/{año}', fontsize=16, fontweight='bold')
        
        ruta_img = os.path.join(self.ruta_salida, f'tasa_resolucion_{mes}_{año}.png')
        plt.savefig(ruta_img, dpi=300, bbox_inches='tight')
        plt.close()
        imagenes.append(ruta_img)
        
        return imagenes
    
    def generar_reporte_excel(self, estadisticas, mes, año, imagenes):
        """
        Genera el reporte final en formato Excel con formato profesional.
        
        Args:
            estadisticas (dict): Estadísticas calculadas
            mes (int): Mes del reporte
            año (int): Año del reporte
            imagenes (list): Lista de rutas a imágenes generadas
        """
        nombre_archivo = f'reporte_indicadores_{mes}_{año}.xlsx'
        ruta_archivo = os.path.join(self.ruta_salida, nombre_archivo)
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Resumen Ejecutivo'
        
        # Configurar estilos
        titulo_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        
        encabezado_font = Font(name='Arial', size=12, bold=True)
        encabezado_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        encabezado_font_blanco = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        
        # Título principal
        ws.merge_cells('A1:E1')
        ws['A1'] = f'REPORTE DE INDICADORES - {mes}/{año}'
        ws['A1'].font = titulo_font
        ws['A1'].fill = titulo_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Fecha de generación
        ws['A3'] = 'Fecha de generación:'
        ws['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws['A3'].font = encabezado_font
        
        # Resumen de indicadores
        ws['A5'] = 'RESUMEN DE INDICADORES'
        ws['A5'].font = encabezado_font
        ws['A5'].fill = encabezado_fill
        ws['A5'].font = encabezado_font_blanco
        ws.merge_cells('A5:B5')
        
        # Datos del resumen
        resumen_datos = [
            ('Total de Incidencias', estadisticas['total_incidencias']),
            ('Incidencias Resueltas', estadisticas['incidencias_resueltas']),
            ('Tasa de Resolución', f"{estadisticas['tasa_resolucion']:.1f}%"),
            ('Tiempo Promedio de Resolución', f"{estadisticas['tiempo_promedio_resolucion']:.1f} horas")
        ]
        
        fila_actual = 6
        for etiqueta, valor in resumen_datos:
            ws[f'A{fila_actual}'] = etiqueta
            ws[f'B{fila_actual}'] = valor
            ws[f'A{fila_actual}'].font = Font(bold=True)
            fila_actual += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        # Guardar archivo
        wb.save(ruta_archivo)
        print(f"Reporte generado exitosamente: {ruta_archivo}")
        
    def ejecutar_reporte_mensual(self, mes, año):
        """
        Ejecuta el proceso completo de generación de reporte mensual.
        
        Args:
            mes (int): Mes del reporte
            año (int): Año del reporte
        """
        print(f"Iniciando generación de reporte para {mes}/{año}...")
        
        # Cargar datos
        df_incidencias = self.cargar_datos_incidencias()
        
        if df_incidencias.empty:
            print("No se encontraron datos para procesar.")
            return
        
        # Generar estadísticas
        estadisticas = self.generar_estadisticas_mensuales(df_incidencias, mes, año)
        
        # Crear visualizaciones
        imagenes = self.crear_visualizaciones(estadisticas, mes, año)
        
        # Generar reporte Excel
        self.generar_reporte_excel(estadisticas, mes, año, imagenes)
        
        print("Proceso completado exitosamente.")


# Ejemplo de uso
if __name__ == "__main__":
    generador = GeneradorReportesIndicadores()
    generador.ejecutar_reporte_mensual(7, 2025)  # Reporte de julio 2025